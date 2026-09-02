from __future__ import annotations

import csv
import json
import mimetypes
import re
import zipfile

from datetime import date, datetime, time
from pathlib import Path, PurePosixPath
from typing import Any, Iterable

from ..config import Settings
from ..db import LedgerDB
from ..telemetry import StageRunner
from ..util import (
    json_dumps,
    new_id,
    utc_now,
)


ZIP_EXTENSIONS = {
    "zip",
}

WORKBOOK_EXTENSIONS = {
    "xlsx",
    "xlsm",
    "xltx",
    "xltm",
    "xls",
    "xlsb",
}

CONTAINER_EXTENSIONS = (
    ZIP_EXTENSIONS
    | WORKBOOK_EXTENSIONS
)

OPENPYXL_EXTENSIONS = {
    "xlsx",
    "xlsm",
    "xltx",
    "xltm",
}

XLRD_EXTENSIONS = {
    "xls",
}

PYXLSB_EXTENSIONS = {
    "xlsb",
}

SKIP_NAMES = {
    ".DS_Store",
    "Thumbs.db",
}

MAX_CONTAINER_DEPTH = 8

CSV_MIME_TYPE = "text/csv"

WORKBOOK_DERIVATION_METHOD = (
    "workbook_sheet_to_csv"
)

WORKBOOK_TRIAGE_MODE = (
    "first_reportable_hit"
)


def _safe_member_path(
    member_name: str,
) -> Path | None:
    #
    # Normalize ZIP member names and block
    # absolute/parent traversal paths.
    #
    pure = PurePosixPath(
        member_name.replace(
            "\\",
            "/",
        )
    )

    if pure.is_absolute():
        return None

    parts = [
        part
        for part in pure.parts
        if part not in {
            "",
            ".",
        }
    ]

    if (
        not parts
        or any(
            part == ".."
            for part in parts
        )
    ):
        return None

    return Path(
        *parts
    )


def _unique_child_path(
    base: Path,
    rel: Path,
) -> Path:
    candidate = (
        base
        / rel
    )

    if not candidate.exists():
        return candidate

    stem = candidate.stem
    suffix = candidate.suffix
    parent = candidate.parent

    n = 2

    while True:
        alt = (
            parent
            / f"{stem}__{n}{suffix}"
        )

        if not alt.exists():
            return alt

        n += 1


def _safe_sheet_filename(
    sheet_name: str,
    sheet_index: int,
) -> str:
    clean = str(
        sheet_name
        or "Sheet"
    ).strip()

    clean = re.sub(
        r'[<>:"/\\|?*\x00-\x1f]',
        "_",
        clean,
    )

    clean = re.sub(
        r"\s+",
        "_",
        clean,
    )

    clean = clean.strip(
        " ._"
    )

    if not clean:
        clean = "Sheet"

    #
    # Keep generated path lengths sane.
    #
    clean = clean[:120]

    return (
        f"Sheet_{sheet_index:04d}"
        f"__{clean}.csv"
    )


def _normalize_extension(
    extension: str | None,
) -> str:
    return (
        str(
            extension
            or ""
        )
        .strip()
        .lower()
        .lstrip(".")
    )


def _is_zip_extension(
    extension: str | None,
) -> bool:
    return (
        _normalize_extension(
            extension
        )
        in ZIP_EXTENSIONS
    )


def _is_workbook_extension(
    extension: str | None,
) -> bool:
    return (
        _normalize_extension(
            extension
        )
        in WORKBOOK_EXTENSIONS
    )


def _is_container_extension(
    extension: str | None,
) -> bool:
    return (
        _normalize_extension(
            extension
        )
        in CONTAINER_EXTENSIONS
    )


def _row_value(
    row,
    name: str,
    default: Any = None,
) -> Any:
    try:
        if (
            hasattr(
                row,
                "keys",
            )
            and name
            not in row.keys()
        ):
            return default

        value = row[name]

        return (
            default
            if value is None
            else value
        )

    except Exception:
        return default


def _cell_to_csv_value(
    value: Any,
) -> Any:
    """
    Normalize spreadsheet cell values for deterministic
    worksheet-derived CSV output.
    """

    if value is None:
        return ""

    if isinstance(
        value,
        datetime,
    ):
        return value.isoformat(
            sep=" "
        )

    if isinstance(
        value,
        date,
    ):
        return value.isoformat()

    if isinstance(
        value,
        time,
    ):
        return value.isoformat()

    if isinstance(
        value,
        bool,
    ):
        return (
            "TRUE"
            if value
            else "FALSE"
        )

    return value


def _trim_trailing_empty_cells(
    row_values: Iterable[Any],
) -> list[Any]:
    values = [
        _cell_to_csv_value(
            value
        )
        for value in row_values
    ]

    last_nonempty = -1

    for index, value in enumerate(
        values
    ):
        if str(
            value
        ).strip():
            last_nonempty = index

    if last_nonempty < 0:
        return []

    return values[
        : last_nonempty + 1
    ]


def _write_sheet_rows_to_csv(
    rows: Iterable[Iterable[Any]],
    csv_path: Path,
) -> dict[str, Any]:
    """
    Stream a worksheet to CSV.

    Internal blank rows are preserved.

    Trailing blank rows and trailing blank columns are
    omitted so heavily formatted spreadsheets do not
    generate enormous empty CSVs.
    """

    csv_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    row_count = 0
    nonblank_row_count = 0
    max_column_count = 0
    pending_blank_rows = 0
    cell_value_count = 0

    #
    # utf-8-sig makes the derivative CSV friendlier to
    # Excel/Cyber Utility Suite while remaining valid UTF-8.
    #
    with csv_path.open(
        "w",
        newline="",
        encoding="utf-8-sig",
    ) as handle:
        writer = csv.writer(
            handle,
            lineterminator="\n",
        )

        for raw_row in rows:
            row_count += 1

            values = (
                _trim_trailing_empty_cells(
                    raw_row
                )
            )

            if not values:
                pending_blank_rows += 1
                continue

            #
            # Preserve internal blank rows but do not flush
            # trailing blank rows at EOF.
            #
            for _ in range(
                pending_blank_rows
            ):
                writer.writerow(
                    []
                )

            pending_blank_rows = 0

            writer.writerow(
                values
            )

            nonblank_row_count += 1

            max_column_count = max(
                max_column_count,
                len(values),
            )

            cell_value_count += sum(
                1
                for value in values
                if str(
                    value
                ).strip()
            )

    is_blank = (
        nonblank_row_count == 0
    )

    if is_blank:
        try:
            csv_path.unlink()
        except FileNotFoundError:
            pass

        size_bytes = 0

    else:
        size_bytes = (
            csv_path
            .stat()
            .st_size
        )

    return {
        "is_blank": is_blank,
        "rows_seen": row_count,
        "nonblank_row_count": (
            nonblank_row_count
        ),
        "max_column_count": (
            max_column_count
        ),
        "nonblank_cell_count": (
            cell_value_count
        ),
        "csv_bytes": size_bytes,
    }


def _openpyxl_visibility(
    worksheet,
) -> str:
    value = str(
        getattr(
            worksheet,
            "sheet_state",
            "visible",
        )
        or "visible"
    ).strip().lower()

    if value in {
        "visible",
        "hidden",
        "veryhidden",
        "very_hidden",
    }:
        if value == "veryhidden":
            return "very_hidden"

        return value

    return "unknown"


def _xls_visibility(
    workbook,
    sheet_index_zero_based: int,
) -> str:
    #
    # xlrd may expose sheet_visibility depending on
    # workbook/version.
    #
    visibility = getattr(
        workbook,
        "sheet_visibility",
        None,
    )

    if not isinstance(
        visibility,
        (
            list,
            tuple,
        ),
    ):
        return "unknown"

    try:
        value = int(
            visibility[
                sheet_index_zero_based
            ]
        )
    except Exception:
        return "unknown"

    return {
        0: "visible",
        1: "hidden",
        2: "very_hidden",
    }.get(
        value,
        "unknown",
    )


def _iter_openpyxl_workbook(
    workbook_path: Path,
):
    try:
        from openpyxl import (
            load_workbook,
        )
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for "
            "XLSX/XLSM/XLTX/XLTM expansion."
        ) from exc

    #
    # data_only=True deliberately exposes cached/display
    # values rather than formula source text for Detection.
    #
    workbook = load_workbook(
        filename=str(
            workbook_path
        ),
        read_only=True,
        data_only=True,
        keep_links=False,
    )

    try:
        for index, worksheet in enumerate(
            workbook.worksheets,
            start=1,
        ):
            yield {
                "sheet_index": index,
                "sheet_name": (
                    worksheet.title
                ),
                "sheet_visibility": (
                    _openpyxl_visibility(
                        worksheet
                    )
                ),
                "rows": (
                    worksheet.iter_rows(
                        values_only=True
                    )
                ),
            }

    finally:
        workbook.close()


def _iter_xls_workbook(
    workbook_path: Path,
):
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError(
            "xlrd is required for "
            "legacy XLS expansion."
        ) from exc

    workbook = xlrd.open_workbook(
        str(
            workbook_path
        ),
        on_demand=True,
    )

    try:
        for sheet_index_zero in range(
            workbook.nsheets
        ):
            worksheet = (
                workbook.sheet_by_index(
                    sheet_index_zero
                )
            )

            def row_iterator(
                sheet=worksheet,
            ):
                for row_index in range(
                    sheet.nrows
                ):
                    yield (
                        sheet.row_values(
                            row_index
                        )
                    )

            yield {
                "sheet_index": (
                    sheet_index_zero
                    + 1
                ),
                "sheet_name": (
                    worksheet.name
                ),
                "sheet_visibility": (
                    _xls_visibility(
                        workbook,
                        sheet_index_zero,
                    )
                ),
                "rows": row_iterator(),
            }

            try:
                workbook.unload_sheet(
                    sheet_index_zero
                )
            except Exception:
                pass

    finally:
        try:
            workbook.release_resources()
        except Exception:
            pass


def _iter_xlsb_workbook(
    workbook_path: Path,
):
    try:
        from pyxlsb import (
            open_workbook,
        )
    except ImportError as exc:
        raise RuntimeError(
            "pyxlsb is required for "
            "XLSB expansion."
        ) from exc

    with open_workbook(
        str(
            workbook_path
        )
    ) as workbook:
        sheet_names = list(
            workbook.sheets
        )

        for index, sheet_name in enumerate(
            sheet_names,
            start=1,
        ):
            with workbook.get_sheet(
                sheet_name
            ) as worksheet:

                def row_iterator(
                    sheet=worksheet,
                ):
                    for row in sheet.rows():
                        yield [
                            getattr(
                                cell,
                                "v",
                                None,
                            )
                            for cell in row
                        ]

                yield {
                    "sheet_index": index,
                    "sheet_name": (
                        sheet_name
                    ),
                    #
                    # pyxlsb does not reliably expose workbook
                    # visibility state through its public API.
                    #
                    "sheet_visibility": (
                        "unknown"
                    ),
                    "rows": row_iterator(),
                }


def _iter_workbook_sheets(
    workbook_path: Path,
    extension: str,
):
    ext = _normalize_extension(
        extension
    )

    if ext in OPENPYXL_EXTENSIONS:
        yield from (
            _iter_openpyxl_workbook(
                workbook_path
            )
        )
        return

    if ext in XLRD_EXTENSIONS:
        yield from (
            _iter_xls_workbook(
                workbook_path
            )
        )
        return

    if ext in PYXLSB_EXTENSIONS:
        yield from (
            _iter_xlsb_workbook(
                workbook_path
            )
        )
        return

    raise RuntimeError(
        "Unsupported workbook extension "
        f"for expansion: {ext}"
    )


def _insert_child_file(
    db: LedgerDB,
    *,
    parent_row,
    parent_file_id: str,
    child_file_id: str,
    child_path: Path,
    logical_path: str,
    extension: str,
    mime_type: str,
    source_bytes: int,
    container_depth: int,
    container_path: str,
    child_stage_status: dict[str, Any] | None = None,
) -> None:
    now = utc_now()

    family_id = str(
        _row_value(
            parent_row,
            "family_id",
            "",
        )
        or parent_file_id
    )

    #
    # The worksheet or ZIP member is both a source-container
    # child and a logical child of the immediate parent.
    #
    db.execute(
        """
        INSERT INTO file_processing_metrics (
            file_id,
            matter_id,
            job_id,
            custodian_id,
            original_path,
            normalized_path,
            extension,
            mime_type,
            source_bytes,
            expanded_bytes,
            is_container,
            is_extracted,
            source_container_file_id,
            parent_file_id,
            family_id,
            container_depth,
            container_path,
            stage_status_json,
            created_at,
            updated_at
        )
        SELECT
            ?,
            matter_id,
            job_id,
            custodian_id,
            ?,
            ?,
            ?,
            ?,
            ?,
            ?,
            0,
            1,
            ?,
            ?,
            CASE
                WHEN coalesce(?, '') <> ''
                    THEN ?
                ELSE family_id
            END,
            ?,
            ?,
            ?,
            ?,
            ?
        FROM file_processing_metrics
        WHERE file_id=?
        """,
        (
            child_file_id,
            str(
                child_path
            ),
            logical_path,
            extension,
            mime_type,
            source_bytes,
            source_bytes,
            parent_file_id,
            parent_file_id,
            family_id,
            family_id,
            container_depth,
            container_path,
            json_dumps(
                child_stage_status
                or {}
            ),
            now,
            now,
            parent_file_id,
        ),
    )


def _queue_child_container(
    queue: list[dict[str, Any]],
    *,
    child_file_id: str,
    child_path: Path,
    logical_path: str,
    extension: str,
    source_bytes: int,
    parent_file_id: str,
    container_depth: int,
    family_id: str = "",
) -> None:
    queue.append(
        {
            "file_id": (
                child_file_id
            ),
            "original_path": (
                str(
                    child_path
                )
            ),
            "normalized_path": (
                logical_path
            ),
            "extension": (
                extension
            ),
            "source_bytes": (
                source_bytes
            ),
            "source_container_file_id": (
                parent_file_id
            ),
            "parent_file_id": (
                parent_file_id
            ),
            "family_id": (
                family_id
            ),
            "container_depth": (
                container_depth
            ),
            "is_container": 0,
        }
    )


def _expand_zip_container(
    *,
    db: LedgerDB,
    row,
    expansion_root: Path,
    queue: list[dict[str, Any]],
) -> dict[str, Any]:
    file_id = str(
        row["file_id"]
    )

    container_depth = int(
        row["container_depth"]
        or 0
    )

    container_path = Path(
        row["original_path"]
    )

    event_status = "completed"
    event_exceptions: list[
        dict[str, Any]
    ] = []

    extracted_file_count = 0
    extracted_bytes = 0
    nested_container_count = 0

    with zipfile.ZipFile(
        container_path
    ) as zf:
        bad = zf.testzip()

        if bad:
            event_status = (
                "completed_with_warnings"
            )

            event_exceptions.append(
                {
                    "warning": (
                        "zip CRC issue first "
                        f"bad member: {bad}"
                    )
                }
            )

        for member in zf.infolist():
            if member.is_dir():
                continue

            safe_rel = (
                _safe_member_path(
                    member.filename
                )
            )

            if (
                safe_rel is None
                or safe_rel.name
                in SKIP_NAMES
            ):
                continue

            child_base = (
                expansion_root
                / file_id
            )

            child_path = (
                _unique_child_path(
                    child_base,
                    safe_rel,
                )
            )

            child_path.parent.mkdir(
                parents=True,
                exist_ok=True,
            )

            with (
                zf.open(
                    member
                ) as src,
                child_path.open(
                    "wb"
                ) as dst,
            ):
                while True:
                    chunk = src.read(
                        1024 * 1024
                    )

                    if not chunk:
                        break

                    dst.write(
                        chunk
                    )

            stat_size = (
                child_path
                .stat()
                .st_size
            )

            ext = (
                child_path
                .suffix
                .lower()
                .lstrip(".")
            )

            mime_type, _ = (
                mimetypes.guess_type(
                    child_path.name
                )
            )

            child_file_id = (
                new_id(
                    "FILE"
                )
            )

            logical_path = (
                f"{row['normalized_path']}"
                f"!/{safe_rel.as_posix()}"
            )

            _insert_child_file(
                db,
                parent_row=row,
                parent_file_id=file_id,
                child_file_id=(
                    child_file_id
                ),
                child_path=(
                    child_path
                ),
                logical_path=(
                    logical_path
                ),
                extension=ext,
                mime_type=(
                    mime_type
                    or "application/octet-stream"
                ),
                source_bytes=(
                    stat_size
                ),
                container_depth=(
                    container_depth
                    + 1
                ),
                container_path=str(
                    row[
                        "normalized_path"
                    ]
                ),
                child_stage_status={
                    "container_expansion": {
                        "status": (
                            "extracted"
                        ),
                        "source_type": (
                            "zip_member"
                        ),
                        "parent_file_id": (
                            file_id
                        ),
                        "parent_normalized_path": (
                            row[
                                "normalized_path"
                            ]
                        ),
                    }
                },
            )

            extracted_file_count += 1
            extracted_bytes += (
                stat_size
            )

            if _is_container_extension(
                ext
            ):
                nested_container_count += 1

                _queue_child_container(
                    queue,
                    child_file_id=(
                        child_file_id
                    ),
                    child_path=(
                        child_path
                    ),
                    logical_path=(
                        logical_path
                    ),
                    extension=ext,
                    source_bytes=(
                        stat_size
                    ),
                    parent_file_id=(
                        file_id
                    ),
                    container_depth=(
                        container_depth
                        + 1
                    ),
                    family_id=str(
                        _row_value(
                            row,
                            "family_id",
                            "",
                        )
                        or ""
                    ),
                )

    return {
        "status": event_status,
        "exceptions": (
            event_exceptions
        ),
        "extracted_file_count": (
            extracted_file_count
        ),
        "extracted_bytes": (
            extracted_bytes
        ),
        "nested_container_count": (
            nested_container_count
        ),
        "container_type": "zip",
        "workbook_manifest": None,
    }


def _expand_workbook_container(
    *,
    db: LedgerDB,
    row,
    expansion_root: Path,
) -> dict[str, Any]:
    file_id = str(
        row["file_id"]
    )

    workbook_path = Path(
        row["original_path"]
    )

    workbook_extension = (
        _normalize_extension(
            row["extension"]
        )
    )

    container_depth = int(
        row["container_depth"]
        or 0
    )

    workbook_output_dir = (
        expansion_root
        / file_id
        / "workbook_sheets"
    )

    workbook_output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook_manifest: dict[
        str,
        Any,
    ] = {
        "schema_version": 1,
        "source_type": (
            "workbook"
        ),
        "workbook_file_id": (
            file_id
        ),
        "workbook_name": (
            workbook_path.name
        ),
        "workbook_normalized_path": (
            row["normalized_path"]
        ),
        "workbook_format": (
            workbook_extension
        ),
        "derivation_method": (
            WORKBOOK_DERIVATION_METHOD
        ),
        "triage_detection_mode": (
            WORKBOOK_TRIAGE_MODE
        ),
        "sheet_count": 0,
        "nonblank_sheet_count": 0,
        "blank_sheet_count": 0,
        "generated_child_count": 0,
        "sheets": [],
    }

    event_exceptions: list[
        dict[str, Any]
    ] = []

    extracted_file_count = 0
    extracted_bytes = 0

    for sheet in _iter_workbook_sheets(
        workbook_path,
        workbook_extension,
    ):
        sheet_index = int(
            sheet[
                "sheet_index"
            ]
        )

        sheet_name = str(
            sheet[
                "sheet_name"
            ]
            or f"Sheet{sheet_index}"
        )

        sheet_visibility = str(
            sheet.get(
                "sheet_visibility"
            )
            or "unknown"
        )

        workbook_manifest[
            "sheet_count"
        ] += 1

        csv_name = (
            _safe_sheet_filename(
                sheet_name,
                sheet_index,
            )
        )

        csv_path = (
            workbook_output_dir
            / csv_name
        )

        try:
            sheet_metrics = (
                _write_sheet_rows_to_csv(
                    sheet["rows"],
                    csv_path,
                )
            )

        except Exception as exc:
            event_exceptions.append(
                {
                    "sheet_index": (
                        sheet_index
                    ),
                    "sheet_name": (
                        sheet_name
                    ),
                    "error": repr(
                        exc
                    ),
                }
            )

            workbook_manifest[
                "sheets"
            ].append(
                {
                    "sheet_index": (
                        sheet_index
                    ),
                    "sheet_name": (
                        sheet_name
                    ),
                    "visibility": (
                        sheet_visibility
                    ),
                    "status": (
                        "failed"
                    ),
                    "error": repr(
                        exc
                    ),
                }
            )

            continue

        if sheet_metrics[
            "is_blank"
        ]:
            workbook_manifest[
                "blank_sheet_count"
            ] += 1

            workbook_manifest[
                "sheets"
            ].append(
                {
                    "sheet_index": (
                        sheet_index
                    ),
                    "sheet_name": (
                        sheet_name
                    ),
                    "visibility": (
                        sheet_visibility
                    ),
                    "status": (
                        "blank"
                    ),
                    "child_file_id": (
                        None
                    ),
                    "child_doc_id": (
                        None
                    ),
                    "rows_seen": (
                        sheet_metrics[
                            "rows_seen"
                        ]
                    ),
                    "nonblank_row_count": (
                        0
                    ),
                    "max_column_count": (
                        0
                    ),
                    "nonblank_cell_count": (
                        0
                    ),
                    "triage_status": (
                        "not_required_blank"
                    ),
                    "counts_complete": (
                        True
                    ),
                }
            )

            continue

        workbook_manifest[
            "nonblank_sheet_count"
        ] += 1

        csv_bytes = int(
            sheet_metrics[
                "csv_bytes"
            ]
        )

        child_file_id = (
            new_id(
                "FILE"
            )
        )

        logical_path = (
            f"{row['normalized_path']}"
            f"!/workbook/"
            f"{csv_name}"
        )

        child_lineage = {
            "workbook_sheet": {
                "schema_version": 1,
                "source_type": (
                    "workbook_sheet"
                ),
                "derived_from": (
                    "workbook_sheet"
                ),
                "derivation_method": (
                    WORKBOOK_DERIVATION_METHOD
                ),
                "original_workbook_file_id": (
                    file_id
                ),
                "original_workbook_name": (
                    workbook_path.name
                ),
                "original_workbook_path": (
                    str(
                        workbook_path
                    )
                ),
                "original_workbook_normalized_path": (
                    row[
                        "normalized_path"
                    ]
                ),
                "workbook_format": (
                    workbook_extension
                ),
                "sheet_name": (
                    sheet_name
                ),
                "sheet_index": (
                    sheet_index
                ),
                "sheet_visibility": (
                    sheet_visibility
                ),
                "sheet_rows_seen": (
                    sheet_metrics[
                        "rows_seen"
                    ]
                ),
                "sheet_nonblank_row_count": (
                    sheet_metrics[
                        "nonblank_row_count"
                    ]
                ),
                "sheet_column_count": (
                    sheet_metrics[
                        "max_column_count"
                    ]
                ),
                "sheet_nonblank_cell_count": (
                    sheet_metrics[
                        "nonblank_cell_count"
                    ]
                ),
                #
                # These values are intentionally seeded
                # now so triage does not require a lineage
                # migration later.
                #
                "triage_status": (
                    "pending"
                ),
                "triage_detection_mode": (
                    WORKBOOK_TRIAGE_MODE
                ),
                "counts_complete": (
                    False
                ),
                "full_iar_status": (
                    "pending_if_hit"
                ),
            },
            "container_expansion": {
                "status": (
                    "extracted"
                ),
                "source_type": (
                    "workbook_sheet"
                ),
                "parent_file_id": (
                    file_id
                ),
            },
        }

        _insert_child_file(
            db,
            parent_row=row,
            parent_file_id=(
                file_id
            ),
            child_file_id=(
                child_file_id
            ),
            child_path=(
                csv_path
            ),
            logical_path=(
                logical_path
            ),
            extension="csv",
            mime_type=(
                CSV_MIME_TYPE
            ),
            source_bytes=(
                csv_bytes
            ),
            container_depth=(
                container_depth
                + 1
            ),
            container_path=str(
                row[
                    "normalized_path"
                ]
            ),
            child_stage_status=(
                child_lineage
            ),
        )

        extracted_file_count += 1
        extracted_bytes += (
            csv_bytes
        )

        workbook_manifest[
            "generated_child_count"
        ] += 1

        workbook_manifest[
            "sheets"
        ].append(
            {
                "sheet_index": (
                    sheet_index
                ),
                "sheet_name": (
                    sheet_name
                ),
                "visibility": (
                    sheet_visibility
                ),
                "status": (
                    "converted"
                ),
                "child_file_id": (
                    child_file_id
                ),
                "child_doc_id": (
                    None
                ),
                "csv_name": (
                    csv_name
                ),
                "csv_path": (
                    str(
                        csv_path
                    )
                ),
                "logical_path": (
                    logical_path
                ),
                "csv_bytes": (
                    csv_bytes
                ),
                "rows_seen": (
                    sheet_metrics[
                        "rows_seen"
                    ]
                ),
                "nonblank_row_count": (
                    sheet_metrics[
                        "nonblank_row_count"
                    ]
                ),
                "max_column_count": (
                    sheet_metrics[
                        "max_column_count"
                    ]
                ),
                "nonblank_cell_count": (
                    sheet_metrics[
                        "nonblank_cell_count"
                    ]
                ),
                "triage_status": (
                    "pending"
                ),
                "triage_detection_mode": (
                    WORKBOOK_TRIAGE_MODE
                ),
                "counts_complete": (
                    False
                ),
            }
        )

    status = "completed"

    if event_exceptions:
        if (
            extracted_file_count > 0
            or workbook_manifest[
                "blank_sheet_count"
            ] > 0
        ):
            status = (
                "completed_with_warnings"
            )
        else:
            status = "failed"

    #
    # A workbook that opened successfully and contains only
    # blank worksheets is still a successful expansion.
    #
    if (
        workbook_manifest[
            "sheet_count"
        ] > 0
        and not event_exceptions
    ):
        status = "completed"

    return {
        "status": status,
        "exceptions": (
            event_exceptions
        ),
        "extracted_file_count": (
            extracted_file_count
        ),
        "extracted_bytes": (
            extracted_bytes
        ),
        "nested_container_count": 0,
        "container_type": (
            "workbook"
        ),
        "workbook_manifest": (
            workbook_manifest
        ),
    }


def run_container_expansion(
    db: LedgerDB,
    settings: Settings,
    job_id: str,
    matter_id: str,
    input_dir: str,
    max_depth: int = MAX_CONTAINER_DEPTH,
) -> None:
    """
    Expand supported ingestion containers locally.

    Supported container families:

      ZIP:
        - .zip
        - nested ZIP/workbook children are recursively queued

      Workbook:
        - .xlsx
        - .xlsm
        - .xltx
        - .xltm
        - .xls
        - .xlsb

    Workbook behavior:

      - original workbook is preserved
      - workbook becomes the parent/container
      - each nonblank worksheet becomes one derived CSV child
      - blank sheets remain in lineage manifest without creating
        unnecessary review documents
      - visible, hidden, and very-hidden sheets are processed
        when visibility metadata is available
      - generated CSV children enter normal downstream
        hash/dedupe/Doc-ID/promotion/detection
      - worksheet lineage is persisted in stage_status_json
      - worksheet triage state is pre-seeded for the later
        first-reportable-hit detection pass

    Successfully expanded containers are marked is_container=1,
    so downstream stages operate on their leaf children.
    """

    root = Path(
        input_dir
    ).resolve()

    expansion_root = (
        root.parent
        / ".apc_expanded"
        / job_id
    )

    expansion_root.mkdir(
        parents=True,
        exist_ok=True,
    )

    placeholders = ",".join(
        "?"
        for _ in sorted(
            CONTAINER_EXTENSIONS
        )
    )

    initial_rows = db.query(
        f"""
        SELECT
            file_id,
            original_path,
            normalized_path,
            extension,
            source_bytes,
            source_container_file_id,
            parent_file_id,
            family_id,
            container_depth,
            is_container
        FROM file_processing_metrics
        WHERE job_id=?
          AND is_container=0
          AND lower(
                coalesce(
                    extension,
                    ''
                )
              ) IN ({placeholders})
        ORDER BY normalized_path
        """,
        (
            job_id,
            *sorted(
                CONTAINER_EXTENSIONS
            ),
        ),
    )

    with StageRunner(
        db,
        settings,
        job_id,
        matter_id,
        "container_expansion",
        "local-container-expander-v2",
    ) as stage:
        stage.metrics.files_in = (
            len(
                initial_rows
            )
        )

        stage.metrics.documents_in = (
            len(
                initial_rows
            )
        )

        stage.metrics.bytes_in = sum(
            int(
                row[
                    "source_bytes"
                ]
                or 0
            )
            for row
            in initial_rows
        )

        queue: list[
            dict[str, Any]
        ] = [
            dict(
                row
            )
            for row
            in initial_rows
        ]

        expanded_container_count = 0
        expanded_zip_count = 0
        expanded_workbook_count = 0

        extracted_file_count = 0
        workbook_sheet_child_count = 0
        workbook_blank_sheet_count = 0

        nested_container_count = 0

        extracted_bytes_total = 0

        max_seen_depth = 0

        exceptions: list[
            dict[str, Any]
        ] = []

        expansion_events = 0

        workbook_manifests: list[
            dict[str, Any]
        ] = []

        while queue:
            row = queue.pop(
                0
            )

            file_id = str(
                row[
                    "file_id"
                ]
            )

            extension = (
                _normalize_extension(
                    row[
                        "extension"
                    ]
                )
            )

            container_depth = int(
                row[
                    "container_depth"
                ]
                or 0
            )

            max_seen_depth = max(
                max_seen_depth,
                container_depth,
            )

            container_path = Path(
                row[
                    "original_path"
                ]
            )

            compressed_bytes = int(
                row[
                    "source_bytes"
                ]
                or 0
            )

            parent_container_file_id = (
                row.get(
                    "source_container_file_id"
                )
                if isinstance(
                    row,
                    dict,
                )
                else _row_value(
                    row,
                    "source_container_file_id",
                    None,
                )
            )

            event_status = (
                "completed"
            )

            event_exceptions: list[
                dict[str, Any]
            ] = []

            event_extracted_files = 0
            event_extracted_bytes = 0
            event_nested_count = 0

            container_type = (
                "unknown"
            )

            workbook_manifest = None

            if (
                container_depth
                >= max_depth
            ):
                event_status = (
                    "skipped_max_depth"
                )

                event_exceptions.append(
                    {
                        "error": (
                            "max container depth "
                            f"{max_depth} reached"
                        )
                    }
                )

            else:
                try:
                    if _is_zip_extension(
                        extension
                    ):
                        result = (
                            _expand_zip_container(
                                db=db,
                                row=row,
                                expansion_root=(
                                    expansion_root
                                ),
                                queue=queue,
                            )
                        )

                    elif _is_workbook_extension(
                        extension
                    ):
                        result = (
                            _expand_workbook_container(
                                db=db,
                                row=row,
                                expansion_root=(
                                    expansion_root
                                ),
                            )
                        )

                    else:
                        raise RuntimeError(
                            "Unsupported container "
                            f"extension: {extension}"
                        )

                    event_status = str(
                        result.get(
                            "status"
                        )
                        or "completed"
                    )

                    event_exceptions = list(
                        result.get(
                            "exceptions"
                        )
                        or []
                    )

                    event_extracted_files = int(
                        result.get(
                            "extracted_file_count"
                        )
                        or 0
                    )

                    event_extracted_bytes = int(
                        result.get(
                            "extracted_bytes"
                        )
                        or 0
                    )

                    event_nested_count = int(
                        result.get(
                            "nested_container_count"
                        )
                        or 0
                    )

                    container_type = str(
                        result.get(
                            "container_type"
                        )
                        or "unknown"
                    )

                    workbook_manifest = (
                        result.get(
                            "workbook_manifest"
                        )
                    )

                except Exception as exc:
                    event_status = (
                        "failed"
                    )

                    event_exceptions.append(
                        {
                            "error": repr(
                                exc
                            )
                        }
                    )

            if (
                event_status
                in {
                    "completed",
                    "completed_with_warnings",
                }
            ):
                stage_payload: dict[
                    str,
                    Any,
                ] = {
                    "container_expansion": {
                        "status": (
                            event_status
                        ),
                        "container_type": (
                            container_type
                        ),
                        "extracted_files": (
                            event_extracted_files
                        ),
                        "extracted_bytes": (
                            event_extracted_bytes
                        ),
                        "nested_container_count": (
                            event_nested_count
                        ),
                    }
                }

                if (
                    workbook_manifest
                    is not None
                ):
                    stage_payload[
                        "workbook_expansion"
                    ] = (
                        workbook_manifest
                    )

                db.execute(
                    """
                    UPDATE file_processing_metrics
                    SET
                        is_container=1,
                        updated_at=?,
                        stage_status_json=json_patch(
                            stage_status_json,
                            ?
                        )
                    WHERE file_id=?
                    """,
                    (
                        utc_now(),
                        json_dumps(
                            stage_payload
                        ),
                        file_id,
                    ),
                )

                expanded_container_count += 1

                if (
                    container_type
                    == "zip"
                ):
                    expanded_zip_count += 1

                elif (
                    container_type
                    == "workbook"
                ):
                    expanded_workbook_count += 1

            else:
                #
                # Failed/skipped containers remain leaf files so
                # they can be exception/remediation coded later.
                #
                exception_record = {
                    "file_id": (
                        file_id
                    ),
                    "path": (
                        str(
                            container_path
                        )
                    ),
                    "status": (
                        event_status
                    ),
                    "details": (
                        event_exceptions
                    ),
                    "extension": (
                        extension
                    ),
                }

                exceptions.append(
                    exception_record
                )

                db.execute(
                    """
                    UPDATE file_processing_metrics
                    SET
                        updated_at=?,
                        exception_json=json_patch(
                            exception_json,
                            ?
                        ),
                        stage_status_json=json_patch(
                            stage_status_json,
                            ?
                        )
                    WHERE file_id=?
                    """,
                    (
                        utc_now(),
                        json_dumps(
                            event_exceptions
                        ),
                        json_dumps(
                            {
                                "container_expansion": {
                                    "status": (
                                        event_status
                                    ),
                                    "container_type": (
                                        container_type
                                    ),
                                }
                            }
                        ),
                        file_id,
                    ),
                )

            db.execute(
                """
                INSERT INTO container_expansion_event (
                    event_id,
                    matter_id,
                    job_id,
                    source_file_id,
                    parent_container_file_id,
                    container_path,
                    original_container_path,
                    container_depth,
                    compressed_bytes,
                    extracted_bytes,
                    extracted_file_count,
                    nested_container_count,
                    status,
                    exception_json,
                    created_at
                )
                VALUES (
                    ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?
                )
                """,
                (
                    new_id(
                        "EXPAND"
                    ),
                    matter_id,
                    job_id,
                    file_id,
                    parent_container_file_id,
                    row[
                        "normalized_path"
                    ],
                    str(
                        container_path
                    ),
                    container_depth,
                    compressed_bytes,
                    event_extracted_bytes,
                    event_extracted_files,
                    event_nested_count,
                    event_status,
                    json_dumps(
                        event_exceptions
                    ),
                    utc_now(),
                ),
            )

            expansion_events += 1

            extracted_file_count += (
                event_extracted_files
            )

            extracted_bytes_total += (
                event_extracted_bytes
            )

            nested_container_count += (
                event_nested_count
            )

            if (
                workbook_manifest
                is not None
            ):
                workbook_manifests.append(
                    workbook_manifest
                )

                workbook_sheet_child_count += int(
                    workbook_manifest.get(
                        "generated_child_count"
                    )
                    or 0
                )

                workbook_blank_sheet_count += int(
                    workbook_manifest.get(
                        "blank_sheet_count"
                    )
                    or 0
                )

        leaf = db.query_one(
            """
            SELECT
                count(*) AS leaf_files,
                coalesce(
                    sum(source_bytes),
                    0
                ) AS leaf_bytes
            FROM file_processing_metrics
            WHERE job_id=?
              AND is_container=0
            """,
            (
                job_id,
            ),
        )

        source = db.query_one(
            """
            SELECT source_bytes
            FROM processing_job
            WHERE job_id=?
            """,
            (
                job_id,
            ),
        )

        source_bytes = (
            int(
                source[
                    "source_bytes"
                ]
                or 0
            )
            if source
            else 0
        )

        expanded_bytes = (
            int(
                leaf[
                    "leaf_bytes"
                ]
                or 0
            )
            if leaf
            else 0
        )

        leaf_files = (
            int(
                leaf[
                    "leaf_files"
                ]
                or 0
            )
            if leaf
            else 0
        )

        expansion_ratio = (
            expanded_bytes
            / source_bytes
            if source_bytes
            else 1.0
        )

        #
        # Blob transaction proxies.
        #
        if extracted_file_count:
            stage.quote_cost(
                "Storage",
                "Blob Write Operations",
                extracted_file_count,
                "operations",
                confidence_note=(
                    "proxy for extracted ZIP members "
                    "and worksheet-derived CSV writes"
                ),
            )

        if expansion_events:
            stage.quote_cost(
                "Storage",
                "Blob Read Operations",
                expansion_events,
                "operations",
                confidence_note=(
                    "proxy for ZIP/workbook "
                    "container reads"
                ),
            )

        stage.metrics.files_in = (
            expansion_events
        )

        stage.metrics.documents_in = (
            expansion_events
        )

        stage.metrics.files_out = (
            leaf_files
        )

        stage.metrics.documents_out = (
            leaf_files
        )

        stage.metrics.bytes_out = (
            expanded_bytes
        )

        stage.metrics.exceptions = (
            len(
                exceptions
            )
        )

        stage.metrics.extra.update(
            {
                "expansion_root": (
                    str(
                        expansion_root
                    )
                ),
                "expanded_container_count": (
                    expanded_container_count
                ),
                "expanded_zip_count": (
                    expanded_zip_count
                ),
                "expanded_workbook_count": (
                    expanded_workbook_count
                ),
                "extracted_file_count": (
                    extracted_file_count
                ),
                "workbook_sheet_child_count": (
                    workbook_sheet_child_count
                ),
                "workbook_blank_sheet_count": (
                    workbook_blank_sheet_count
                ),
                "nested_container_count": (
                    nested_container_count
                ),
                "extracted_bytes_total": (
                    extracted_bytes_total
                ),
                "expanded_leaf_file_count": (
                    leaf_files
                ),
                "expanded_leaf_bytes": (
                    expanded_bytes
                ),
                "expansion_ratio": (
                    expansion_ratio
                ),
                "max_container_depth": (
                    max_seen_depth
                ),
                "container_exceptions": (
                    exceptions[
                        :50
                    ]
                ),
                "supported_containers": (
                    sorted(
                        CONTAINER_EXTENSIONS
                    )
                ),
                "supported_workbooks": (
                    sorted(
                        WORKBOOK_EXTENSIONS
                    )
                ),
                "workbook_manifest_count": (
                    len(
                        workbook_manifests
                    )
                ),
                "workbook_manifests": (
                    workbook_manifests[
                        :25
                    ]
                ),
            }
        )

        db.execute(
            """
            UPDATE processing_job
            SET
                compressed_source_bytes=?,
                expanded_bytes=?,
                expanded_file_count=?,
                container_file_count=?,
                extracted_file_count=?,
                container_exception_count=?,
                max_container_depth=?,
                expansion_ratio=?,
                processed_bytes=?
            WHERE job_id=?
            """,
            (
                source_bytes,
                expanded_bytes,
                leaf_files,
                expanded_container_count,
                extracted_file_count,
                len(
                    exceptions
                ),
                max_seen_depth,
                expansion_ratio,
                expanded_bytes,
                job_id,
            ),
        )